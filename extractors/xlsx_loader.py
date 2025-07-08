from openpyxl import load_workbook
from io import BytesIO

def load_xlsx_text(file_bytes: bytes) -> str:
    wb = load_workbook(filename=BytesIO(file_bytes), data_only=True)
    text = []
    for sheet in wb:
        for row in sheet.iter_rows(values_only=True):
            line = [str(cell) for cell in row if cell is not None]
            if line:
                text.append(" ".join(line))
    return "\n".join(text)